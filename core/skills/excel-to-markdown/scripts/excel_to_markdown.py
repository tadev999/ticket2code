#!/usr/bin/env python3
"""Convert Excel/CSV files to markdown tables.

This script is intentionally dependency-light:
- .xlsx: openpyxl
- .xls: xlrd (if installed)
- .csv: stdlib csv
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import pathlib
import sys
from typing import Iterable, List, Sequence, Tuple


def col_name(index: int) -> str:
    name = ""
    n = index
    while n > 0:
        n, rem = divmod(n - 1, 26)
        name = chr(65 + rem) + name
    return name


def to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    return str(value).strip()


def sanitize_cell(value: str) -> str:
    return value.replace("\n", "<br>").replace("|", "\\|")


def normalize_rows(rows: Sequence[Sequence[object]], max_rows: int, max_cols: int) -> Tuple[List[List[str]], bool, bool]:
    limited_rows = list(rows[:max_rows])
    row_truncated = len(rows) > max_rows

    max_width = 0
    for row in limited_rows:
        if len(row) > max_width:
            max_width = len(row)
    if max_width == 0:
        max_width = 1

    width = min(max_width, max_cols)
    col_truncated = max_width > max_cols

    normalized: List[List[str]] = []
    for row in limited_rows:
        values = [to_text(v) for v in row[:width]]
        if len(values) < width:
            values.extend([""] * (width - len(values)))
        normalized.append(values)

    if not normalized:
        normalized = [[""] * width]

    return normalized, row_truncated, col_truncated


def build_header(first_row: Sequence[str]) -> List[str]:
    non_empty_count = sum(1 for cell in first_row if cell)
    if non_empty_count == 0:
        return [f"Column {col_name(i + 1)}" for i in range(len(first_row))]
    return [cell if cell else f"Column {col_name(i + 1)}" for i, cell in enumerate(first_row)]


def markdown_table(headers: Sequence[str], rows: Sequence[Sequence[str]]) -> str:
    header_line = "| " + " | ".join(sanitize_cell(h) for h in headers) + " |"
    sep_line = "| " + " | ".join("---" for _ in headers) + " |"
    row_lines = ["| " + " | ".join(sanitize_cell(c) for c in row) + " |" for row in rows]
    return "\n".join([header_line, sep_line] + row_lines)


def parse_csv(path: pathlib.Path, encoding: str) -> List[Tuple[str, List[List[object]]]]:
    with path.open("r", encoding=encoding, newline="") as f:
        reader = csv.reader(f)
        rows = [list(r) for r in reader]
    return [(path.stem, rows)]


def parse_xlsx(path: pathlib.Path) -> List[Tuple[str, List[List[object]]]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        raise RuntimeError("Missing dependency 'openpyxl'. Install with: pip install openpyxl") from exc

    wb = load_workbook(filename=path, data_only=True, read_only=True)
    sheets: List[Tuple[str, List[List[object]]]] = []
    for ws in wb.worksheets:
        rows: List[List[object]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        sheets.append((ws.title, rows))
    return sheets


def parse_xls(path: pathlib.Path) -> List[Tuple[str, List[List[object]]]]:
    try:
        import xlrd  # type: ignore
    except ImportError as exc:
        raise RuntimeError("Missing dependency 'xlrd'. Install with: pip install xlrd") from exc

    wb = xlrd.open_workbook(path.as_posix())
    sheets: List[Tuple[str, List[List[object]]]] = []
    for sheet in wb.sheets():
        rows: List[List[object]] = []
        for r in range(sheet.nrows):
            rows.append(sheet.row_values(r))
        sheets.append((sheet.name, rows))
    return sheets


def render_markdown(
    source_file: pathlib.Path,
    sheets: Iterable[Tuple[str, List[List[object]]]],
    max_rows: int,
    max_cols: int,
) -> str:
    blocks: List[str] = [f"# Spreadsheet Extract: {source_file.name}", ""]

    for sheet_name, raw_rows in sheets:
        normalized, row_truncated, col_truncated = normalize_rows(raw_rows, max_rows=max_rows, max_cols=max_cols)
        headers = build_header(normalized[0])
        data_rows = normalized[1:] if len(normalized) > 1 else []

        blocks.append(f"## Sheet: {sheet_name}")
        blocks.append(f"- Source: {source_file.name}")
        blocks.append(f"- Rows (rendered): {len(data_rows)}")
        blocks.append(f"- Columns (rendered): {len(headers)}")
        if row_truncated or col_truncated:
            blocks.append(
                "- Truncation: "
                f"rows>{max_rows}: {'yes' if row_truncated else 'no'}, "
                f"cols>{max_cols}: {'yes' if col_truncated else 'no'}"
            )
        blocks.append("")

        blocks.append(markdown_table(headers, data_rows))
        blocks.append("")

    return "\n".join(blocks).rstrip() + "\n"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Convert Excel/CSV files to markdown")
    parser.add_argument("--input", required=True, help="Path to .xlsx/.xls/.csv")
    parser.add_argument("--output", required=True, help="Output markdown path")
    parser.add_argument("--max-rows", type=int, default=500, help="Max rows per sheet (default: 500)")
    parser.add_argument("--max-cols", type=int, default=30, help="Max columns per sheet (default: 30)")
    parser.add_argument("--encoding", default="utf-8", help="CSV encoding (default: utf-8)")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    input_path = pathlib.Path(args.input).expanduser().resolve()
    output_path = pathlib.Path(args.output).expanduser().resolve()

    if not input_path.exists():
        print(f"ERROR: Input file not found: {input_path}", file=sys.stderr)
        return 1

    ext = input_path.suffix.lower()
    try:
        if ext == ".csv":
            sheets = parse_csv(input_path, args.encoding)
        elif ext == ".xlsx":
            sheets = parse_xlsx(input_path)
        elif ext == ".xls":
            sheets = parse_xls(input_path)
        else:
            print(f"ERROR: Unsupported file type: {ext}", file=sys.stderr)
            return 1

        markdown = render_markdown(
            source_file=input_path,
            sheets=sheets,
            max_rows=max(1, args.max_rows),
            max_cols=max(1, args.max_cols),
        )

        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(markdown, encoding="utf-8")

        sheet_count = len(list(sheets)) if isinstance(sheets, list) else 0
        print(f"Converted: {input_path}")
        print(f"Output: {output_path}")
        print(f"Sheets: {sheet_count}")
        return 0
    except Exception as exc:  # pragma: no cover
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
